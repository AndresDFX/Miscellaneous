import pandas as pd
import numpy as np
import openpyxl
import requests
from openpyxl import load_workbook
import urllib3
from urllib3.exceptions import HTTPError as BaseHTTPError
import sys
import re

def readFile(file, sheet, column_name):
    df = pd.read_excel(file, sheet_name=sheet)
    column_URL = df[column_name].values
    last_column = len(df.columns[-1]) - 2
    return column_URL, last_column


def es_correo(correo):
    expresion_regular = r"(?:[a-z0-9!#$%&'*+/=?^_`{|}~-]+(?:\.[a-z0-9!#$%&'*+/=?^_`{|}~-]+)*|\"(?:[\x01-\x08\x0b\x0c\x0e-\x1f\x21\x23-\x5b\x5d-\x7f]|\\[\x01-\x09\x0b\x0c\x0e-\x7f])*\")@(?:(?:[a-z0-9](?:[a-z0-9-]*[a-z0-9])?\.)+[a-z0-9](?:[a-z0-9-]*[a-z0-9])?|\[(?:(?:(2(5[0-5]|[0-4][0-9])|1[0-9][0-9]|[1-9]?[0-9]))\.){3}(?:(2(5[0-5]|[0-4][0-9])|1[0-9][0-9]|[1-9]?[0-9])|[a-z0-9-]*[a-z0-9]:(?:[\x01-\x08\x0b\x0c\x0e-\x1f\x21-\x5a\x53-\x7f]|\\[\x01-\x09\x0b\x0c\x0e-\x7f])+)\])"

    return re.match(expresion_regular, correo) is not None

#TEST CASES
#print(es_correo("direccionacademica@medised.edu.co"))

def url_check(column_url, SSL):
       #Description

    """Array return - check to see if the site exists.
        This function takes a url as input and then it requests the site
        get content and then it checks the response to see if
        it's less than 400.
    """
    new_column = [0] * len(column_url)

    for i, url in enumerate(column_url):
        try:


            if (es_correo(url.strip()) == False):
                session = requests.Session()
                session.max_redirects = 15 # Optimize time (default 30)

                URLparser = parser(str(url.strip()))
                site_ping = session.get(URLparser, verify=SSL,  headers={'user-agent': 'My app'}) #Add header with pages bad constructed

                if site_ping.status_code < 400 or site_ping.status_code == 406: # 406: Status code not aceptable
                    #print("True")
                    new_column[i] = 'True'
                else:
                    new_column[i] = 'False'
                    #print("False")

            else:
                new_column[i] = 'False'

        except requests.exceptions.SSLError:
                new_column[i] = 'SSL Error'

        except:
                new_column[i] = 'False'

    return new_column


def parser(url):
    if url[:7].strip() == "http://" or url[:8].strip() == "https://":
        return url
    else:
        return "http://"+url

#TEST CASES
#print(parser("http://www.escueladebellezafantasia.amawebs.com"))
#print(parser("www.poliagro@edu.co"))
#print(parser("PENDIENTE"))
#print(url_check2("www.google.com"))
#print(url_check2(parser("www.google.com")))
#print(url_check2("https://www.google.com/"))
#print(url_check2("http://www.escueladebellezafantasia.amawebs.com"))


def url_check2(url):
    try:
        session = requests.Session()
        session.max_redirects = 10 # Optimize time (default 20)
        site_ping = session.get(parser(str(url.strip())), verify=False)

        if site_ping.status_code < 400 or site_ping.status_code == 406:
            return True
        else:
            return False
    except Exception as e:
        print(str(e))

#TEST CASES
#print(url_check2("http://www.existenodominio.com/")) #Not exists domain
#print(url_check2("https://institutotecnisistemas.com/")) #Redirect site (406)
#print(url_check2("PENDIENTE")) # Word
#print(url_check2("http://www.escueladebellezafantasia.amawebs.com")) # Contains http
#print(url_check2("http://corporaciontecnicactc.jimdo.com/")) #The site is not load
#print(url_check2("www.poliagro@edu.co")) #Is not web page
#print(url_check2("direccionacademica@medised.edu.co")) #SSL not active


def append_df_to_excel(filename, df, sheet_name, startrow,
                       truncate_sheet, startcol):


    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, header=["Estado"], index=False , startrow=startrow, startcol=startcol)

    # save the workbook
    writer.save()

# Callback Function
def main(file, sheet, column_name):

    data, startcol = readFile(file, sheet, column_name)
    data = url_check(data, True)
    new_df = pd.DataFrame(data)
    append_df_to_excel(file, new_df, sheet, 0, False, startcol)


# Main function
if __name__ == "__main__":
    main(sys.argv[1],sys.argv[2],sys.argv[3])
    print("Finish verification")